# -*- coding: utf-8 -*-
"""
Adaptado de: Copia de Extractor datos superficie OPAV.ipynb
Logica de negocio sin modificar; solo se quito lo especifico de Colab
(drive.mount, loop de carpetas) y se agrego el wrapper procesar_archivo()
con la interfaz estandar (input_file, output_dir) -> [rutas_generadas].
"""

import os
import re
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuración global
# ---------------------------------------------------------------------------
EMPRESA = "(OPAV)"
ESPECIE = "Hatchery"
PROCESO = "Environment"
PAIS = "COLOMBIA"
LABORATORIO = "Lab BIOS (NO USAR)"
SUBTANALISIS = "Recuento ambientes y superficies"
EDAD = ""
GRANJA = "La Esperanza"

# Microorganismos de CONTEO (UFC/pl)
MICROORGANISMOS_CONTEO = {
    "mesofilos": ("Mesófilos", "UFC/pl"),
    "hongos": ("Hongos", "UFC/pl"),
}

# Microorganismos Ausencia / Presencia
MICROORGANISMOS_AP = {
    "aspergillus_fumigatus": "Aspergillus fumigatus",
    "otros_asper": "Aspergillus spp",
    "e_coli": "Escherichia coli",
    "pseudomonas_sp": "Pseudomonas spp",
    "coliformes": "Coliformes",
}
UNIDAD_AP = ""

# Diccionario para Normalización del Tipo de Muestra (Columna final 'Muestra')
MUESTRAS_MAPPING = {
    "superficie": "Superficie",
    "ambiente": "Ambiente",
    "frotis": "Frotis",
}

# Diccionario de casos específicos puntuales para ETAPA
AREAS_SALAS_ESPECIFICAS = {
    "Bandeja Eq. Granja": "Bandeja Eq. Granja",
    "Bandeja Eq. G": "Bandeja Eq. Granja",
    "Cuarto cartón": "Cuarto cartón",
    "Cuarto cartó": "Cuarto cartón",
    "Laboratorio": "Sala vacuna",
    "Laboratorio Planta": "Sala vacuna",
    "Agua del Destilador (caliente)": "Sala vacuna",
    "Agua destilada*": "Sala vacuna",
    "Destilada laboratorio": "Sala vacuna",
    "Destilador laboratorio vacuna": "Sala vacuna",
    "Laboratorio destilada": "Sala vacuna",
    "Laboratorio sin destilar": "Sala vacuna",
    "Vacuna": "Sala vacuna",
    "Máquina lavado de cubetas": "Sala de lavado",
    "Maquina lavado de cubetas": "Sala de lavado",
}


# ---------------------------------------------------------------------------
# Funciones de Limpieza y Normalización
# ---------------------------------------------------------------------------
def normalizar(texto):
    if not texto:
        return ""
    texto = str(texto).lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = texto.replace(" de ", " ")
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def _definir_etapa(valor_muestra):
    if not valor_muestra:
        return "Revisar"

    texto_str = str(valor_muestra).strip()
    texto_norm = normalizar(texto_str)

    # 1. Reglas dinámicas por prefijos
    if texto_norm.startswith("almacen"):
        return "Almacén"

    if (
        texto_norm.startswith("caja plastica")
        or texto_norm.startswith("lavado bande")
        or texto_norm.startswith("lavado canas")
        or texto_norm.startswith("maquina lavado")
        or "lavado cubetas" in texto_norm
    ):
        return "Sala de lavado"

    if (
        texto_norm.startswith("ducto")
        or texto_norm.startswith("mesa clasific")
        or texto_norm.startswith("zona")
        or "cuarto frio" in texto_norm
    ):
        return "Cuarto frío"

    # Inclusión de Plenum Inc para Sala de Incubación
    if texto_norm.startswith("incubadora") or texto_norm.startswith("plenum inc"):
        return "Sala de incubación"

    if texto_norm.startswith("nacedora") or texto_norm.startswith("plenum nac"):
        return "Sala de nacimiento"

    if (
        texto_norm.startswith("mesa vacuna")
        or texto_norm.startswith("cuarto vacuna")
        or texto_norm.startswith("zona vacuna")
    ):
        return "Sala vacuna"

    if texto_norm.startswith("mesa transf") or texto_norm.startswith("mesas transf"):
        return "Sala transferencia"

    if texto_norm.startswith("salon despac"):
        return "Sala de despacho"

    if texto_norm.startswith("salon sexaje") or texto_norm.startswith("zona sexaje"):
        return "Zona sexaje y vacunacion"

    # 2. Vehículos (Placas de 3 letras + 3 números o nombres de vehículos/camiones)
    vehiculos_conocidos = ["el progreso", "la caridad", "la esperanza", "la fe"]
    if (
        re.match(r"^[a-zA-Z]{3}\s*\d{3}$", texto_str)
        or any(texto_norm.startswith(v) for v in vehiculos_conocidos)
    ):
        return "Vehiculos"

    # 3. Búsqueda en casos puntuales del diccionario
    for orig, etapa in AREAS_SALAS_ESPECIFICAS.items():
        if normalizar(orig) == texto_norm:
            return etapa

    return "Revisar"


def _normalizar_muestra(tipo_muestra):
    if not tipo_muestra:
        return "Revisar"

    muestra_str = str(tipo_muestra).strip()
    muestra_norm = normalizar(muestra_str)

    for orig, norm in MUESTRAS_MAPPING.items():
        if normalizar(orig) == muestra_norm:
            return norm

    return muestra_str if muestra_str else "Revisar"


def normalizar_resultado_ap(valor):
    if valor is None or str(valor).strip() == "":
        return ""
    val_str = str(valor).strip().lower()
    if val_str in ["1", "1.0", "presencia", "presente", "positivo"] or "presente" in val_str or "presencia" in val_str:
        return "presente"
    if val_str in ["0", "0.0", "ausencia", "ausente", "negativo"] or "ausente" in val_str or "ausencia" in val_str:
        return "ausente"
    return str(valor).strip()


def _extraer_municipio_departamento(ciudad):
    if not isinstance(ciudad, str):
        return None, None

    partes = ciudad.split("-")
    if len(partes) < 2:
        return ciudad.strip(), None

    return partes[0].strip().upper(), partes[1].strip().upper()


# ---------------------------------------------------------------------------
# Extracción de Datos Excel
# ---------------------------------------------------------------------------
def _buscar_valor_por_etiqueta(ws, etiqueta, max_row=15, max_col=20):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            valor = ws.cell(row=r, column=c).value
            if isinstance(valor, str) and etiqueta.lower() in valor.lower():
                for c2 in range(c + 1, max_col + 1):
                    v2 = ws.cell(row=r, column=c2).value
                    if v2 not in (None, ""):
                        return v2
    return None


def extraer_metadatos(ws):
    return {
        "centro_produccion": _buscar_valor_por_etiqueta(
            ws, "CENTRO DE PRODUCCION"
        ),
        "direccion": _buscar_valor_por_etiqueta(ws, "DIRECCION"),
        "ciudad": _buscar_valor_por_etiqueta(ws, "CIUDAD"),
    }


COLUMNAS = {
    1: "n_solicitud",
    2: "codigo_muestra",
    3: "fecha_toma",
    4: "fecha_recepcion",
    5: "fecha_analisis",
    6: "tipo_muestra",     # Columna 'TIPO DE MUESTRA' en el Excel original
    8: "muestra_col",      # Columna 'Muestra' en el Excel original
    9: "area_sala_col",   # Columna 'Área / Sala' en el Excel original
    10: "mesofilos",
    11: "hongos",
    12: "aspergillus_fumigatus",
    13: "otros_asper",
    14: "e_coli",
    15: "pseudomonas_sp",
    16: "coliformes",
    17: "observaciones",
    18: "fecha_emision_informe",
}


def extraer_registros(ws):
    registros = []
    max_row = ws.max_row
    r = 1
    while r <= max_row:
        val_a = ws.cell(row=r, column=1).value
        es_encabezado = (
            isinstance(val_a, str) and "N° SOLICITUD" in val_a.upper()
        )
        if es_encabezado:
            fila_datos = r + 3
            while fila_datos <= max_row:
                codigo = ws.cell(row=fila_datos, column=2).value
                n_solicitud = ws.cell(row=fila_datos, column=1).value
                if codigo is None and n_solicitud is None:
                    break
                if codigo is not None:
                    registro = {
                        nombre: ws.cell(row=fila_datos, column=col).value
                        for col, nombre in COLUMNAS.items()
                    }
                    registros.append(registro)
                fila_datos += 1
            r = fila_datos
        else:
            r += 1
    return registros


# ---------------------------------------------------------------------------
# Orquestación y Procesamiento
# ---------------------------------------------------------------------------
def _procesar_archivo_ambiente_superficie(ruta_entrada, ruta_salida):
    wb = load_workbook(ruta_entrada, data_only=True)

    todos_los_registros = []

    # Auditar valores marcados como "Revisar"
    muestras_no_normalizadas = set()
    tipos_muestra_no_normalizados = set()

    for ws in wb.worksheets:
        metadatos = extraer_metadatos(ws)
        registros = extraer_registros(ws)

        for reg in registros:
            reg.update(metadatos)
            reg["Empresa"] = f"Planta de Incubación {EMPRESA}"
            reg["Granja"] = GRANJA
            reg["Especie"] = ESPECIE
            reg["proceso"] = PROCESO
            reg["Pais"] = PAIS

            municipio, departamento = _extraer_municipio_departamento(
                metadatos.get("ciudad")
            )
            reg["Municipio"] = municipio
            reg["Departamento"] = departamento
            reg["idlote"] = ""
            reg["Edad"] = EDAD

            # Extracción de valores del Excel original
            muestra_excel = reg.pop("muestra_col", "") or ""          # Columna 8 del Excel
            area_sala_excel = reg.pop("area_sala_col", "") or ""      # Columna 9 del Excel
            tipo_muestra_excel = reg.get("tipo_muestra", "") or ""   # Columna 6 del Excel

            # Galpon se asigna con el área/sala original
            reg["Galpon"] = area_sala_excel

            # DEFINICIÓN DE ETAPA: Evaluando 'Muestra' (columna 8) del Excel original
            etapa_mapeada = _definir_etapa(muestra_excel)
            reg["Etapa"] = etapa_mapeada
            if etapa_mapeada == "Revisar" and muestra_excel:
                muestras_no_normalizadas.add(str(muestra_excel).strip())

            # MUESTRA (columna final): Extrae y normaliza 'TIPO DE MUESTRA' (columna 6)
            tipo_muestra_mapeado = _normalizar_muestra(tipo_muestra_excel)
            reg["Muestra"] = tipo_muestra_mapeado
            if tipo_muestra_mapeado == "Revisar" and tipo_muestra_excel:
                tipos_muestra_no_normalizados.add(str(tipo_muestra_excel).strip())

            # CODIGO MUESTRA: Concatenación de "CODIGO DE MUESTRA - Muestra (columna 8)"
            cod_m = str(reg.get('codigo_muestra', '')).strip()
            muestra_orig_str = str(muestra_excel).strip()
            reg["Codigo Muestra"] = f"{cod_m} - {muestra_orig_str}"

            reg["Fecha Toma Muestra"] = reg.get("fecha_toma")
            reg["Fecha Recepcion Laboratorio"] = reg.get("fecha_recepcion")
            reg["Fecha Analisis"] = reg.get("fecha_analisis")
            reg["Laboratorio"] = LABORATORIO
            reg["Subtipo de Analisis"] = SUBTANALISIS

            # Extraer microorganismos
            valores_conteo = {
                nombre_bonito: (reg.pop(campo, None), unidad)
                for campo, (
                    nombre_bonito,
                    unidad,
                ) in MICROORGANISMOS_CONTEO.items()
            }
            valores_ap = {
                nombre_bonito: reg.pop(campo, None)
                for campo, nombre_bonito in MICROORGANISMOS_AP.items()
            }

            # Filas Conteos (Mesófilos, Hongos)
            for nombre_micro, (valor, unidad) in valores_conteo.items():
                if valor is None or str(valor).strip() == "":
                    continue
                fila = reg.copy()
                fila["Microorganismo"] = nombre_micro
                fila["Resultado"] = valor
                fila["Unidad De Medida"] = unidad
                todos_los_registros.append(fila)

            # Filas Ausencia / Presencia (Aspergillus, E. coli, Pseudomonas, Coliformes)
            for nombre_micro, valor in valores_ap.items():
                if valor is None or str(valor).strip() == "":
                    continue
                fila = reg.copy()
                fila["Microorganismo"] = nombre_micro
                fila["Resultado"] = normalizar_resultado_ap(valor)
                fila["Unidad De Medida"] = UNIDAD_AP
                todos_los_registros.append(fila)

    columnas_orden = [
        "Empresa",
        "Granja",
        "Especie",
        "Etapa",
        "Pais",
        "Departamento",
        "Municipio",
        "Galpon",
        "idlote",
        "Edad",
        "Muestra",
        "Codigo Muestra",
        "Fecha Toma Muestra",
        "Laboratorio",
        "Fecha Recepcion Laboratorio",
        "Fecha Analisis",
        "Subtipo de Analisis",
        "Microorganismo",
        "Resultado",
        "Unidad De Medida",
    ]

    df = pd.DataFrame(todos_los_registros)
    if not df.empty:
        df = df[columnas_orden]

    df.to_excel(ruta_salida, index=False, sheet_name="Datos limpios")
    _formatear_salida(ruta_salida, df)

    # ---------------------------------------------------------------------------
    # REPORTE DE AUDITORÍA EN CONSOLA
    # ---------------------------------------------------------------------------
    print("\n" + "=" * 60)
    print(f"📊 REPORTE DE PROCESAMIENTO: {os.path.basename(ruta_entrada)}")
    print(f"   Total registros extraídos: {len(df)}")

    hay_pendientes = False

    if muestras_no_normalizadas:
        hay_pendientes = True
        print(
            f"\n   ⚠️ VALORES EN 'MUESTRA' NO NORMALIZADOS PARA ETAPA ({len(muestras_no_normalizadas)}):"
        )
        print("      (Fueron marcados como 'Revisar' en la columna Etapa del Excel)")
        for item in sorted(muestras_no_normalizadas):
            print(f"      - '{item}'")

    if tipos_muestra_no_normalizados:
        hay_pendientes = True
        print(
            f"\n   ⚠️ TIPOS DE MUESTRA NO NORMALIZADOS ({len(tipos_muestra_no_normalizados)}):"
        )
        print("      (Fueron marcados como 'Revisar' en la columna Muestra del Excel)")
        for item in sorted(tipos_muestra_no_normalizados):
            print(f"      - '{item}'")

    if not hay_pendientes:
        print(
            "\n   🎉 ¡ÉXITO! Todos los campos fueron normalizados correctamente."
        )

    print("=" * 60 + "\n")

    return df


def _formatear_salida(ruta_salida, df):
    wb = load_workbook(ruta_salida)
    ws = wb.active
    fuente = "Arial"

    for row in ws.iter_rows():
        for celda in row:
            celda.font = Font(name=fuente, size=10)

    for celda in ws[1]:
        celda.font = Font(name=fuente, size=10, bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center")

    columnas_fecha = {
        "Fecha Toma Muestra",
        "Fecha Recepcion Laboratorio",
        "Fecha Analisis",
    }
    for idx, nombre_col in enumerate(df.columns, start=1):
        letra = get_column_letter(idx)
        ancho = (
            max(12, min(40, df[nombre_col].astype(str).map(len).max() + 2))
            if not df.empty
            else 15
        )
        ws.column_dimensions[letra].width = ancho
        if nombre_col in columnas_fecha:
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=idx).number_format = "DD/MM/YYYY"

    wb.save(ruta_salida)


# ---------------------------------------------------------------------------
# INTERFAZ ESTANDAR PARA EL ROUTER
# ---------------------------------------------------------------------------
def procesar_archivo(input_file, output_dir):
    """
    input_file : ruta local al Excel de Ambiente y superficie descargado
                 de Gmail.
    output_dir : carpeta donde debe quedar el Excel limpio generado.
    Devuelve la lista de rutas generadas.
    """
    os.makedirs(output_dir, exist_ok=True)
    nombre_archivo = os.path.basename(input_file)
    ruta_salida = os.path.join(output_dir, f"limpio_{nombre_archivo}")

    _procesar_archivo_ambiente_superficie(input_file, ruta_salida)

    return [ruta_salida]