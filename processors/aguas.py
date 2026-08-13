# -*- coding: utf-8 -*-
"""
Adaptado de: Copia de Extractor datos aguas OPAV.ipynb
Logica de negocio sin modificar; solo se quito lo especifico de Colab
(drive.mount, loop de carpetas) y se agrego el wrapper procesar_archivo()
con la interfaz estandar (input_file, output_dir) -> [rutas_generadas].
"""

import os
import re
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuración global
# ---------------------------------------------------------------------------
EMPRESA = "(OPAV)"
ESPECIE = "Hatchery"
PROCESO = "Environment"
PAIS = "COLOMBIA"
LABORATORIO = "Lab BIOS (NO USAR)"
SUBTANALISIS = "Microbiológico de agua"
EDAD = 1
GRANJA = "La Esperanza"

UNIDAD_CONTEO = "UFC/100 ml"
MICROORGANISMOS_CONTEO = {
    "mesofilos": "Mesófilos",
    "coliformes_totales": "Coliformes Totales",
    "coliformes_fecales": "Coliformes Fecales",
}

AREAS_SALAS = {
    "Laboratorio": "Sala vacuna",
    "Agua del Destilador (caliente)": "Sala vacuna",
    "Agua destilada*": "Sala vacuna",
    "Clasificación cuarto frío": "Cuarto frío",
    "Cuarto vacuna": "Sala vacuna",
    "Destilada laboratorio": "Sala vacuna",
    "Destilador laboratorio vacuna": "Sala vacuna",
    "Incubadora A1": "Sala de incubación",
    "Incubadora sala A1": "Sala de incubación",
    "Incubadora sala C": "Sala de incubación",
    "Incubadoras sala A2": "Sala de incubación",
    "Incubadoras sala C": "Sala de incubación",
    "Laboratorio destilada": "Sala vacuna",
    "Laboratorio Planta": "Sala vacuna",
    "Laboratorio sin destilar": "Sala vacuna",
    "Lavado bandejas": "Sala de lavado",
    "Lavado bandejas transferencia": "Sala de lavado",
    "Lavado canastillas": "Sala de lavado",
    "Lavado canastillas zona sucia": "Sala de lavado",
    "Lavado canatillas": "Sala de lavado",
    "Lavado cubetas": "Cuarto frío",
    "Lavado cubetas clasificación": "Cuarto frío",
    "Lavado cubetas cuarto frio": "Cuarto frío",
    "Nacedora A1": "Sala de nacimiento",
    "Nacedora B": "Sala de nacimiento",
    "Nacedora sala A2": "Sala de nacimiento",
    "Nacedora sala B": "Sala de nacimiento",
    "Nacedora sala C": "Sala de nacimiento",
    "Nacedoras sala A1": "Sala de nacimiento",
    "Nacedoras sala A2": "Sala de nacimiento",
    "Nacedoras sala B": "Sala de nacimiento",
    "Nacedoras sala C": "Sala de nacimiento",
    "Zona de vacuna": "Zona sexaje y vacunación",
    "Zona sexaje": "Zona sexaje y vacunación",
    "Zona vacunación": "Zona sexaje y vacunación",
    "Zona clasificación cuarto frío": "Cuarto frío",
    "Incubadoras sala A1": "Sala de incubación",
    "Incubadora sala A1": "Sala de incubación",
    "Incubadora sala A2": "Sala de incubación",
    "Incubadora sala B": "Sala de incubación",
    "Incubadora sala C": "Sala de incubación",
    "Zona cuarto frio clasificación": "Cuarto frío",
    "Zona cuarto frio": "Cuarto frío",
    "Zona clasificacion cuarto frio": "Cuarto frío",
    "Entrada chiller tanque": "Otras áreas",
    "Salida chiller tanque": "Otras áreas"
}


# ---------------------------------------------------------------------------
# Funciones de Limpieza y Normalización
# ---------------------------------------------------------------------------
def normalizar(texto):
    if not texto:
        return ""
    texto = str(texto).lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = texto.replace(" de ", " ")
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def _definir_etapa(galpon):
    if not galpon:
        return "Revisar"

    galpon_norm = normalizar(galpon)

    for area, etapa in AREAS_SALAS.items():
        if normalizar(area) == galpon_norm:
            return etapa

    return "Revisar"


def _extraer_granja(centro_produccion):
    if not isinstance(centro_produccion, str):
        return centro_produccion
    return re.sub(
        r"(?i)^planta de incubaci[oó]n\s*", "", centro_produccion
    ).strip()


def _extraer_municipio_departamento(ciudad):
    if not isinstance(ciudad, str):
        return None, None

    partes = ciudad.split("-")
    if len(partes) < 2:
        return ciudad.strip(), None

    return partes[0].strip().upper(), partes[1].strip().upper()


# ---------------------------------------------------------------------------
# Extracción de Datos Excel
# ---------------------------------------------------------------------------
def _buscar_valor_por_etiqueta(ws, etiqueta, max_row=15, max_col=28):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            valor = ws.cell(row=r, column=c).value
            if isinstance(valor, str) and etiqueta.lower() in valor.lower():
                for c2 in range(c + 1, max_col + 1):
                    v2 = ws.cell(row=r, column=c2).value
                    if v2 not in (None, ""):
                        return v2
    return None


def extraer_metadatos(ws):
    return {
        "centro_produccion": _buscar_valor_por_etiqueta(
            ws, "CENTRO DE PRODUCCION"
        ),
        "direccion": _buscar_valor_por_etiqueta(ws, "DIRECCION"),
        "ciudad": _buscar_valor_por_etiqueta(ws, "CIUDAD"),
    }


COLUMNAS = {
    1: "n_solicitud",
    2: "codigo_muestra",
    3: "fecha_toma",
    5: "fecha_recepcion",
    7: "fecha_analisis",
    11: "tipo_muestra",
    15: "muestra",
    21: "mesofilos",
    22: "coliformes_totales",
    23: "coliformes_fecales",
    24: "observaciones",
    26: "fecha_emision_informe",
}


def extraer_registros(ws):
    registros = []
    max_row = ws.max_row
    r = 1
    while r <= max_row:
        val_a = ws.cell(row=r, column=1).value
        es_encabezado = (
            isinstance(val_a, str) and "N° SOLICITUD" in val_a.upper()
        )
        if es_encabezado:
            fila_datos = r + 3
            while fila_datos <= max_row:
                codigo = ws.cell(row=fila_datos, column=2).value
                n_solicitud = ws.cell(row=fila_datos, column=1).value
                if codigo is None and n_solicitud is None:
                    break
                if codigo is not None:
                    registro = {
                        nombre: ws.cell(row=fila_datos, column=col).value
                        for col, nombre in COLUMNAS.items()
                    }
                    registros.append(registro)
                fila_datos += 1
            r = fila_datos
        else:
            r += 1
    return registros


# ---------------------------------------------------------------------------
# Orquestación y Procesamiento
# ---------------------------------------------------------------------------
def _procesar_archivo_aguas(ruta_entrada, ruta_salida):
    wb = load_workbook(ruta_entrada, data_only=True)

    todos_los_registros = []
    no_normalizados = set()

    for ws in wb.worksheets:
        metadatos = extraer_metadatos(ws)
        registros = extraer_registros(ws)

        for reg in registros:
            reg.update(metadatos)
            reg["Empresa"] = f"Planta de Incubación {EMPRESA}"
            reg["Granja"] = GRANJA
            reg["Especie"] = ESPECIE
            reg["proceso"] = PROCESO
            reg["Pais"] = PAIS

            municipio, departamento = _extraer_municipio_departamento(
                metadatos.get("ciudad")
            )
            reg["Municipio"] = municipio
            reg["Departamento"] = departamento
            reg["idlote"] = ""
            reg["Edad"] = EDAD

            galpon_original = reg.pop("muestra", "") or ""
            reg["Galpon"] = galpon_original

            # Normalización de Etapa
            etapa_mapeada = _definir_etapa(galpon_original)
            reg["Etapa"] = etapa_mapeada

            if etapa_mapeada == "Revisar" and galpon_original:
                no_normalizados.add(str(galpon_original).strip())

            reg["Codigo Muestra"] = (
                f"{reg.get('codigo_muestra', '')}-{reg.get('Galpon', '')}"
            )
            reg["Muestra"] = reg.get("tipo_muestra", "")
            reg["Fecha Toma Muestra"] = reg.get("fecha_toma")
            reg["Fecha Recepción Laboratorio"] = reg.get("fecha_recepcion")
            reg["Fecha Analisis"] = reg.get("fecha_analisis")
            reg["Tratamiento de Agua"] = "Tratada"
            reg["Ubicacion Muestra"] = "true"
            reg["Laboratorio"] = LABORATORIO
            reg["Subtipo de Analisis"] = SUBTANALISIS

            valores_conteo = {
                nombre_bonito: reg.pop(campo, None)
                for campo, nombre_bonito in MICROORGANISMOS_CONTEO.items()
            }

            for nombre_micro, valor in valores_conteo.items():
                if valor in (None, ""):
                    continue
                fila = reg.copy()
                fila["Microorganismo"] = nombre_micro
                fila["Resultado"] = valor
                fila["Unidad De Medida"] = UNIDAD_CONTEO
                todos_los_registros.append(fila)

    columnas_orden = [
        "Empresa",
        "Granja",
        "Especie",
        "Etapa",
        "Pais",
        "Departamento",
        "Municipio",
        "Galpon",
        "idlote",
        "Edad",
        "Muestra",
        "Codigo Muestra",
        "Fecha Toma Muestra",
        "Tratamiento de Agua",
        "Ubicacion Muestra",
        "Laboratorio",
        "Fecha Recepción Laboratorio",
        "Fecha Analisis",
        "Subtipo de Analisis",
        "Microorganismo",
        "Resultado",
        "Unidad De Medida",
    ]

    df = pd.DataFrame(todos_los_registros)
    if not df.empty:
        df = df[columnas_orden]

    df.to_excel(ruta_salida, index=False, sheet_name="Datos limpios")
    _formatear_salida(ruta_salida, df)

    # --- REPORTE DE AUDITORÍA ---
    print(f"\n✅ Procesado: {os.path.basename(ruta_entrada)}")
    print(f"   Total filas extraídas: {len(df)}")

    if no_normalizados:
        print(
            f"   ⚠️ ATENCIÓN: Se encontraron {len(no_normalizados)} área(s) no normalizada(s) (Marcadas como 'Revisar'):"
        )
        for area in sorted(no_normalizados):
            print(f"      - '{area}'")
    else:
        print("   🎉 Todas las áreas/salas fueron normalizadas correctamente.")

    return df


def _formatear_salida(ruta_salida, df):
    wb = load_workbook(ruta_salida)
    ws = wb.active
    fuente = "Arial"

    for row in ws.iter_rows():
        for celda in row:
            celda.font = Font(name=fuente, size=10)

    for celda in ws[1]:
        celda.font = Font(name=fuente, size=10, bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center")

    columnas_fecha = {
        "Fecha Toma Muestra",
        "Fecha Recepción Laboratorio",
        "Fecha Analisis",
    }
    for idx, nombre_col in enumerate(df.columns, start=1):
        letra = get_column_letter(idx)
        ancho = (
            max(12, min(40, df[nombre_col].astype(str).map(len).max() + 2))
            if not df.empty
            else 15
        )
        ws.column_dimensions[letra].width = ancho
        if nombre_col in columnas_fecha:
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=idx).number_format = "DD/MM/YYYY"

    wb.save(ruta_salida)


# ---------------------------------------------------------------------------
# INTERFAZ ESTANDAR PARA EL ROUTER
# ---------------------------------------------------------------------------
def procesar_archivo(input_file, output_dir):
    """
    input_file : ruta local al Excel de Aguas descargado de Gmail.
    output_dir : carpeta donde debe quedar el Excel limpio generado.
    Devuelve la lista de rutas generadas.
    """
    os.makedirs(output_dir, exist_ok=True)
    nombre_archivo = os.path.basename(input_file)
    ruta_salida = os.path.join(output_dir, f"limpio_{nombre_archivo}")

    _procesar_archivo_aguas(input_file, ruta_salida)

    return [ruta_salida]