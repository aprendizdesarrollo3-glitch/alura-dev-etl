# -*- coding: utf-8 -*-
"""
Adaptado de: Copia de Extractor datos vacunas OPAV.ipynb
Logica de negocio sin modificar; solo se quito lo especifico de Colab
(drive.mount, loop de carpetas) y se agrego el wrapper procesar_archivo()
con la interfaz estandar (input_file, output_dir) -> [rutas_generadas].
"""

import os
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONSTANTES
# ---------------------------------------------------------------------------
EMPRESA = "(OPAV)"
ESPECIE = "Hatchery"
PROCESO = "Incubation"
PAIS = "Colombia"
LABORATORIO = "Lab BIOS (NO USAR)"
SUBTANALISIS = "Análisis cualitativo de superficies"

MICROORGANISMOS = {
    "e_coli": "Escherichia coli",
    "pseudomonas_sp": "Pseudomonas spp",
    "coliformes": "Coliformes",
}
UNIDAD_MEDIDA = ""

FUENTE_NORMAL = Font(name="Arial", size=10)
FUENTE_NEGRITA = Font(name="Arial", size=10, bold=True)
ALINEACION_HEADER = Alignment(horizontal="center", vertical="center")

no_normalizados = []  # se acumula a lo largo de TODA la corrida del proceso (por diseño)

# ---------------------------------------------------------------------------
# METADATOS
# ---------------------------------------------------------------------------
def _extraer_municipio_departamento(ciudad):
    if not isinstance(ciudad, str):
        return None, None
    partes = ciudad.split("-")
    if len(partes) < 2:
        return ciudad.strip(), None
    return partes[0].strip(), partes[1].strip()

def _buscar_valor_por_etiqueta(ws, etiqueta, max_row=15, max_col=16):
    etiqueta_limpia = etiqueta.lower().strip()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            valor = ws.cell(row=r, column=c).value
            if isinstance(valor, str) and etiqueta_limpia in valor.lower():
                for c2 in range(c + 1, max_col + 1):
                    v2 = ws.cell(row=r, column=c2).value
                    if v2 not in (None, ""):
                        val_str = str(v2).strip()
                        if val_str.endswith(":"):
                            continue
                        return val_str
    return None

def extraer_metadatos(ws):
    return {
        "centro_produccion": _buscar_valor_por_etiqueta(ws, "Planta de Incubación"),
        "ciudad": _buscar_valor_por_etiqueta(ws, "Ciudad"),
    }

# ---------------------------------------------------------------------------
# EXTRAER REGISTROS TABULARES
# ---------------------------------------------------------------------------
COLUMNAS = {
    1: "n_solicitud", 2: "Codigo Muestra", 3: "Fecha Toma Muestra",
    4: "Fecha Recepcion Laboratorio", 5: "Fecha Analisis", 6: "Muestra",
    8: "Galpon", 10: "Etapa", 11: "e_coli", 12: "pseudomonas_sp",
    13: "coliformes", 14: "observaciones", 15: "fecha_emision_informe",
}

def extraer_registros(ws):
    registros = []
    max_row = ws.max_row
    r = 1
    while r <= max_row:
        val_a = ws.cell(row=r, column=1).value
        if isinstance(val_a, str) and "N° SOLICITUD" in val_a.upper():
            fila_datos = r + 3
            while fila_datos <= max_row:
                codigo = ws.cell(row=fila_datos, column=2).value
                n_solicitud = ws.cell(row=fila_datos, column=1).value
                if codigo is None and n_solicitud is None:
                    break
                if codigo is not None:
                    registros.append({
                        nombre: ws.cell(row=fila_datos, column=col).value
                        for col, nombre in COLUMNAS.items()
                    })
                fila_datos += 1
            r = fila_datos
        else:
            r += 1
    return registros

# ---------------------------------------------------------------------------
# NORMALIZACIÓN
# ---------------------------------------------------------------------------
MAPEO_TIPO_MUESTRA = {
    "diluente vacuna": "Diluente vacuna", "diluente": "Diluente vacuna", "in line": "In Line",
    "manos clasificadores": "Manos Clasificadores", "manos vacunadores": "Manos Vacunadores",
    "máquinas inyectadoras": "Máquinas inyectadoras", "maquinas inyectadoras": "Máquinas inyectadoras",
    "máquinas inyectoras": "Máquinas inyectoras", "maquinas inyectoras": "Máquinas inyectoras",
}

def _normalizar_tipo_muestra(tipo_excel, nombre_archivo="", n_solicitud="", codigo_muestra=""):
    if not isinstance(tipo_excel, str) or not tipo_excel.strip():
        return "revisar"
    clave = tipo_excel.strip().lower()
    if clave in MAPEO_TIPO_MUESTRA:
        return MAPEO_TIPO_MUESTRA[clave]
    no_normalizados.append({
        "archivo": nombre_archivo, "solicitud": n_solicitud,
        "codigo": codigo_muestra, "tipo_original": tipo_excel,
    })
    return "revisar"

# Palabras clave para detectar resultados cualitativos en texto
_TEXTO_AUSENTE = ("negativo", "ausencia", "ausente")
_TEXTO_PRESENTE = ("positivo", "presencia", "presente")

def _es_columna_ap(valores):
    """
    Decide si una columna (microorganismo) es de Ausencia/Presencia.
    Es A/P si:
      - contiene algún texto tipo ausencia/presencia/negativo/positivo, o
      - todos sus valores numéricos son exclusivamente 0 y 1.
    Si tiene números distintos de 0/1 (conteos), NO es A/P.
    """
    tiene_texto_ap = False
    numericos = []
    for v in valores:
        if v is None or str(v).strip() == "":
            continue
        s = str(v).strip().lower()
        if any(t in s for t in _TEXTO_AUSENTE + _TEXTO_PRESENTE):
            tiene_texto_ap = True
            continue
        try:
            numericos.append(float(s))
        except ValueError:
            pass  # texto no reconocido: se ignora para la decisión
    if tiene_texto_ap:
        return True
    if numericos and all(n in (0.0, 1.0) for n in numericos):
        return True
    return False

def _normalizar_resultado(valor, es_ap):
    """
    Normaliza un resultado individual.
      - Texto ausencia/negativo  -> 'Ausente'
      - Texto presencia/positivo -> 'Presente'
      - Numérico en columna A/P   -> 0='Ausente', 1='Presente'
      - Numérico en columna conteo-> se mantiene el número (int si es entero)
    """
    if valor is None or str(valor).strip() == "":
        return ""
    s = str(valor).strip().lower()
    if any(t in s for t in _TEXTO_AUSENTE):
        return "Ausente"
    if any(t in s for t in _TEXTO_PRESENTE):
        return "Presente"
    # Intentar interpretar como número
    try:
        num = float(s)
    except ValueError:
        return str(valor).strip()  # texto desconocido: se deja tal cual
    if es_ap:
        return "Ausente" if num == 0 else "Presente"
    # Columna de conteo: mantener el número
    return int(num) if num.is_integer() else num

# ---------------------------------------------------------------------------
# PROCESAMIENTO PRINCIPAL
# ---------------------------------------------------------------------------
COLUMNAS_ORDEN = [
    "Empresa", "Granja", "Especie", "Etapa", "Pais", "Departamento", "Municipio",
    "Galpon", "idlote", "Edad", "Muestra", "Codigo Muestra", "Fecha Toma Muestra",
    "Laboratorio", "Fecha Recepcion Laboratorio", "Fecha Analisis",
    "Subtipo de Analisis", "Microorganismo", "Resultado", "unidad de medida",
]
COLUMNAS_FECHA = {"Fecha Toma Muestra", "Fecha Recepcion Laboratorio", "Fecha Analisis"}

def _procesar_archivo_vacunacion(ruta_entrada, ruta_salida):
    # read_only=True: no vamos a modificar el archivo de entrada, así que
    # ahorra memoria significativamente en archivos grandes o con imágenes
    wb = load_workbook(ruta_entrada, data_only=True, read_only=True)

    registros_base = []  # 1ª pasada: registros con metadatos y micros aún crudos
    nombre_archivo = os.path.basename(ruta_entrada)

    for ws in wb.worksheets:
        metadatos = extraer_metadatos(ws)
        registros = extraer_registros(ws)
        municipio, departamento = _extraer_municipio_departamento(metadatos.get("ciudad"))

        for reg in registros:
            reg.update(metadatos)
            reg["Empresa"] = f"Planta de Incubación {EMPRESA}"
            reg["Granja"] = "La Esperanza"
            reg["Especie"] = ESPECIE
            reg["Proceso"] = PROCESO
            reg["Pais"] = PAIS
            reg["Municipio"] = municipio
            reg["Departamento"] = departamento
            reg["idlote"] = ""
            reg["Edad"] = ""
            reg["Codigo Muestra"] = f"{reg.get('Codigo Muestra', '')}-{reg.get('Galpon', '')}"
            reg["Laboratorio"] = LABORATORIO
            reg["Subtipo de Analisis"] = SUBTANALISIS
            reg["Muestra"] = _normalizar_tipo_muestra(
                reg.get("Muestra", ""),
                nombre_archivo=nombre_archivo,
                n_solicitud=reg.get("n_solicitud", ""),
                codigo_muestra=reg.get("Codigo Muestra", ""),
            )
            registros_base.append(reg)

    wb.close()  # libera el archivo de entrada explícitamente

    # --- Decidir A/P por microorganismo (a nivel de columna) ---
    es_ap = {}
    for clave in MICROORGANISMOS:
        valores_col = [r.get(clave) for r in registros_base]
        es_ap[clave] = _es_columna_ap(valores_col)

    # --- 2ª pasada: melt (una fila por microorganismo con valor) ---
    todos_los_registros = []
    for reg in registros_base:
        for clave, nombre_micro in MICROORGANISMOS.items():
            valor = reg.get(clave)
            if valor in (None, "") or (isinstance(valor, str) and not valor.strip()):
                continue
            fila = {k: v for k, v in reg.items() if k not in MICROORGANISMOS}
            fila["Microorganismo"] = nombre_micro
            fila["Resultado"] = _normalizar_resultado(valor, es_ap[clave])
            fila["unidad de medida"] = UNIDAD_MEDIDA
            todos_los_registros.append(fila)

    df = pd.DataFrame(todos_los_registros)
    if not df.empty:
        df = df[COLUMNAS_ORDEN]

    _escribir_salida(ruta_salida, df)
    print(f"Listo: {len(df)} registros extraídos -> {ruta_salida}")
    print(f"   Columnas A/P detectadas: "
          f"{[MICROORGANISMOS[k] for k, v in es_ap.items() if v] or 'ninguna'}")
    return df

def _escribir_salida(ruta_salida, df):
    """Escribe y formatea en una sola pasada (evita escribir con pandas
    y luego reabrir el archivo con openpyxl para darle formato)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos limpios"

    # encabezados
    for c_idx, col in enumerate(df.columns, start=1):
        celda = ws.cell(row=1, column=c_idx, value=col)
        celda.font = FUENTE_NEGRITA
        celda.alignment = ALINEACION_HEADER

    # datos
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, valor in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=valor).font = FUENTE_NORMAL

    # anchos de columna + formato de fecha
    if not df.empty:
        anchos = df.astype(str).apply(lambda c: c.str.len().max())
    for idx, nombre_col in enumerate(df.columns, start=1):
        letra = get_column_letter(idx)
        ancho = max(12, min(40, int(anchos[nombre_col]) + 2)) if not df.empty else 15
        ws.column_dimensions[letra].width = ancho
        if nombre_col in COLUMNAS_FECHA:
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=idx).number_format = "DD/MM/YYYY"

    wb.save(ruta_salida)


# ---------------------------------------------------------------------------
# INTERFAZ ESTANDAR PARA EL ROUTER
# ---------------------------------------------------------------------------
def procesar_archivo(input_file, output_dir):
    """
    input_file : ruta local al Excel de Vacunación descargado de Gmail.
    output_dir : carpeta donde debe quedar el Excel limpio generado.
    Devuelve la lista de rutas generadas (para que local_test.py y,
    mas adelante, la Lambda sepan que subir a S3).
    """
    os.makedirs(output_dir, exist_ok=True)
    nombre_archivo = os.path.basename(input_file)
    ruta_salida = os.path.join(output_dir, f"limpio_{nombre_archivo}")

    _procesar_archivo_vacunacion(input_file, ruta_salida)

    return [ruta_salida]